import openpyxl
from decimal import Decimal
from utils import DataClass

cal = """Ссылка на товар
Артикул
Название
Цена
Описание
Ссылки на изображения через запятую
Описание
Все характеристики с сохранением их структуры
Название селлера
Ссылка на селлера
Размеры товара через запятую 
Остатки по товару (число)
Рейтинг
Количество отзывов""".split('\n')
print(len(cal))

file_path = 'data_save.xlsx'
file_path_filter = 'data_save_filter.xlsx'
try:
    workbook = openpyxl.load_workbook(file_path)
    workbookfilter = openpyxl.load_workbook(file_path_filter)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbookfilter = openpyxl.Workbook()

sheet = workbook.active
sheetfilter= workbookfilter.active

if sheet.max_row == 1 and sheet['A1'].value is None:
    sheet.append(cal)

if sheetfilter.max_row == 1 and sheetfilter['A1'].value is None:
    sheetfilter.append(cal)



def save(date:dict):
    all = []
    #Ещё один костыль. Зато работает) Здесь мы получаем все наши списки карточек которые спарсили )
    for i in date:
        for j in i['card']:
            for z in j:
                    all.append(z)
    print(all)
    for info in all:
        new_data = [
            str(info.urlCard),
            str(info.article),
            str(info.name),
            str(info.price),
            str(info.description),
            str(info.imgs),
            str(info.description),
            str(info.opts),
            str(info.supplierName),
            str(info.url_supplier),
            str(info.sizesName),
            str(info.qty),
            str(info.rating),
            str(info.feedBack),
        ]

        #доп файл. Важно на прямую float нельзя сравнивать, тем более в javascript) Используем Decimal

        # Цена
        #Страна РФ
        country = False
        if (Decimal(4.5) <= Decimal(str(info.rating))) and (str(info.opts).upper().find("россия".upper())) and (Decimal(str(info.price)) < Decimal(10000)):
            #доп фильтр и файл)
            sheetfilter.append(new_data)

        sheet.append(new_data)
    workbook.save(file_path)
    workbookfilter.save(file_path_filter)




